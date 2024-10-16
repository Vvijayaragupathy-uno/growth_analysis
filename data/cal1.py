import math
import json
import glob
import openpyxl
import matplotlib.pyplot as plt
from openpyxl.chart import LineChart, Reference

def calculate_max_manhattan_distance(filepath):
    with open(filepath) as f:
        data = json.load(f)
    contours = data['contours']
    max_manhattan_distance = 0
    for i in range(len(contours)):
        for j in range(i + 1, len(contours)):
            point1 = contours[i]
            point2 = contours[j]
            manhattan_distance = abs(point2[0] - point1[0]) + abs(point2[1] - point1[1])
            if manhattan_distance > max_manhattan_distance:
                max_manhattan_distance = manhattan_distance
    return max_manhattan_distance

def calculate_percentage_difference(value1, value2):
    if value1 == 0:
        return float('inf')
    return ((value2 - value1) / abs(value1)) * 100

# Get a list of JSON files in the directory
json_files = sorted(glob.glob('segmentation_points/*.json'))

# Create a new Excel workbook and select the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Manhattan Distance Analysis"

# Write headers
sheet['A1'] = "Current File"
sheet['B1'] = "Next File"
sheet['C1'] = "Current Max Distance"
sheet['D1'] = "Next Max Distance"
sheet['E1'] = "Percentage Difference"

# Initialize list to store percentage differences for plotting
percentage_differences = []

# Iterate through the JSON files to calculate percentage differences
for i in range(len(json_files) - 1):
    current_file = json_files[i]
    next_file = json_files[i + 1]
    
    current_max_distance = calculate_max_manhattan_distance(current_file)
    next_max_distance = calculate_max_manhattan_distance(next_file)
    
    percentage_difference = calculate_percentage_difference(current_max_distance, next_max_distance)
    
    # Write data to Excel
    row = i + 2  # Start from row 2
    sheet.cell(row=row, column=1, value=current_file)
    sheet.cell(row=row, column=2, value=next_file)
    sheet.cell(row=row, column=3, value=current_max_distance)
    sheet.cell(row=row, column=4, value=next_max_distance)
    sheet.cell(row=row, column=5, value=percentage_difference)
    
    percentage_differences.append(percentage_difference)

# Save the Excel file
excel_filename = "manhattan_distance_analysis.xlsx"
workbook.save(excel_filename)
print(f"Excel file '{excel_filename}' has been created and saved.")

# Create a line plot of percentage differences
plt.figure(figsize=(10, 6))
plt.plot(range(1, len(percentage_differences) + 1), percentage_differences, marker='o')
plt.title("Percentage Difference in Max Manhattan Distance")
plt.xlabel("Comparison Number")
plt.ylabel("Percentage Difference (%)")
plt.grid(True)

# Save the plot as an image file
plot_filename = "percentage_difference_plot.png"
plt.savefig(plot_filename)
print(f"Plot has been saved as '{plot_filename}'.")

# Display the plot (optional, comment out if running in an environment without display)
plt.show()